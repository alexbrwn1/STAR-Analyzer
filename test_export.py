#!/usr/bin/env python3
"""
Test script for STAR Analyzer Excel export.
"""

import sys
from pathlib import Path

# Add src to path
src_path = Path(__file__).parent / 'src'
sys.path.insert(0, str(src_path))

from core.parser import parse_medpc_file
from core.file_discovery import discover_medpc_files
from core.exporters import ExcelExporter
from core.data_models import Cohort


def test_excel_export():
    """Test Excel export."""
    print("Testing Excel export...")

    # Parse all test files
    folder = Path("Mock Cohort 1 Data by Day")
    files = discover_medpc_files(folder, recursive=True)

    sessions = []
    for f in files:
        try:
            session = parse_medpc_file(f)
            sessions.append(session)
        except Exception as e:
            print(f"Error parsing {f.name}: {e}")

    print(f"Parsed {len(sessions)} sessions")

    # Create cohort
    cohort = Cohort(
        name="Mock Cohort 1",
        sessions=sessions,
        source_path=folder,
    )

    # Export
    output_path = Path("test_export.xlsx")
    exporter = ExcelExporter()
    exporter.export_cohort(cohort, output_path, include_timestamps=True)

    print(f"Exported to: {output_path}")
    print(f"File size: {output_path.stat().st_size / 1024:.1f} KB")

    # Verify file exists and can be opened
    try:
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        print(f"Sheets: {wb.sheetnames}")
        summary = wb['Summary']
        print(f"Summary rows: {summary.max_row}")
        wb.close()
        print("Excel file validated successfully!")
        return True
    except Exception as e:
        print(f"Error validating Excel file: {e}")
        return False


if __name__ == "__main__":
    success = test_excel_export()
    sys.exit(0 if success else 1)
