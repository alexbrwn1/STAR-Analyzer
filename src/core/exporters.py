"""
Export functionality for STAR Analyzer.
Supports Excel (.xlsx) export of session data.
"""

from datetime import datetime
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from .data_models import MedPCSession, Cohort


class ExcelExporter:
    """Export session data to Excel format."""

    # Column definitions for summary sheet
    SUMMARY_COLUMNS = [
        ('Subject', 12),
        ('Date', 12),
        ('Time', 10),
        ('Experiment', 25),
        ('Group', 20),
        ('Box', 6),
        ('Active Presses', 14),
        ('Inactive Presses', 16),
        ('Discrimination %', 14),
        ('Reinforcers', 12),
        ('Licks', 10),
        ('Session Time (s)', 16),
        ('MSN', 30),
        ('File', 40),
    ]

    def __init__(self):
        self.header_font = Font(bold=True)
        self.header_fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

    def export_sessions(
        self,
        sessions: List[MedPCSession],
        output_path: Path,
        cohort_name: Optional[str] = None,
        include_timestamps: bool = False,
    ) -> None:
        """
        Export sessions to Excel file.

        Args:
            sessions: List of sessions to export
            output_path: Path for output .xlsx file
            cohort_name: Optional name for the export
            include_timestamps: If True, add sheets with timestamp arrays
        """
        wb = Workbook()

        # Create summary sheet
        self._create_summary_sheet(wb, sessions, cohort_name)

        # Optionally add timestamp sheets
        if include_timestamps and sessions:
            self._create_timestamp_sheets(wb, sessions)

        # Save workbook
        wb.save(output_path)

    def export_cohort(
        self,
        cohort: Cohort,
        output_path: Path,
        include_timestamps: bool = False,
    ) -> None:
        """
        Export a cohort to Excel file.

        Args:
            cohort: Cohort to export
            output_path: Path for output .xlsx file
            include_timestamps: If True, add sheets with timestamp arrays
        """
        self.export_sessions(
            cohort.sessions,
            output_path,
            cohort_name=cohort.name,
            include_timestamps=include_timestamps,
        )

    def _create_summary_sheet(
        self,
        wb: Workbook,
        sessions: List[MedPCSession],
        cohort_name: Optional[str] = None,
    ) -> None:
        """Create the main summary sheet with session data."""
        ws = wb.active
        ws.title = 'Summary'

        # Add title row if cohort name provided
        start_row = 1
        if cohort_name:
            ws.cell(row=1, column=1, value=f'STAR Analyzer Export: {cohort_name}')
            ws.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws.cell(row=2, column=1, value=f'Exported: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
            ws.cell(row=3, column=1, value=f'Sessions: {len(sessions)}')
            start_row = 5

        # Add headers
        header_row = start_row
        for col_idx, (header, width) in enumerate(self.SUMMARY_COLUMNS, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Add data rows
        for row_idx, session in enumerate(sessions, start=header_row + 1):
            self._write_session_row(ws, row_idx, session)

        # Freeze header row
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    def _write_session_row(self, ws, row_idx: int, session: MedPCSession) -> None:
        """Write a single session row to the worksheet."""
        meta = session.metadata
        scalars = session.scalars

        # Calculate total licks
        total_licks = scalars.lick_onsets + scalars.lick_offsets

        # Calculate discrimination index (active / total presses * 100)
        total_presses = scalars.active_lever_presses + scalars.inactive_lever_presses
        if total_presses > 0:
            discrimination = round((scalars.active_lever_presses / total_presses) * 100, 1)
        else:
            discrimination = 0.0

        data = [
            meta.subject,
            meta.start_date.strftime('%Y-%m-%d'),
            meta.start_time.strftime('%H:%M:%S'),
            meta.experiment,
            meta.group,
            meta.box,
            scalars.active_lever_presses,
            scalars.inactive_lever_presses,
            discrimination,
            scalars.reinforcers,
            total_licks,
            round(scalars.session_time, 2),
            meta.msn,
            str(meta.file_path.name),
        ]

        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = self.thin_border
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    def _create_timestamp_sheets(self, wb: Workbook, sessions: List[MedPCSession]) -> None:
        """Create sheets with timestamp arrays for each session."""
        # Group by subject for organization
        subjects = sorted(set(s.subject for s in sessions))

        for subject in subjects:
            subject_sessions = [s for s in sessions if s.subject == subject]

            for session in subject_sessions:
                # Create sheet name (Excel has 31 char limit)
                date_str = session.metadata.start_date.strftime('%m%d')
                time_str = session.metadata.start_time.strftime('%H%M')
                sheet_name = f'S{subject}_{date_str}_{time_str}'[:31]

                # Ensure unique sheet name
                base_name = sheet_name
                counter = 1
                while sheet_name in wb.sheetnames:
                    sheet_name = f'{base_name[:28]}_{counter}'
                    counter += 1

                ws = wb.create_sheet(title=sheet_name)
                self._write_timestamp_sheet(ws, session)

    def _write_timestamp_sheet(self, ws, session: MedPCSession) -> None:
        """Write timestamp arrays for a single session."""
        # Write metadata header
        ws.cell(row=1, column=1, value='Subject:').font = Font(bold=True)
        ws.cell(row=1, column=2, value=session.metadata.subject)
        ws.cell(row=2, column=1, value='Date:').font = Font(bold=True)
        ws.cell(row=2, column=2, value=session.metadata.start_date.strftime('%Y-%m-%d'))
        ws.cell(row=3, column=1, value='Time:').font = Font(bold=True)
        ws.cell(row=3, column=2, value=session.metadata.start_time.strftime('%H:%M:%S'))

        # Define arrays to export
        arrays = [
            ('Active Lever (J)', session.timestamps.active_lever_timestamps),
            ('Inactive Lever (K)', session.timestamps.inactive_lever_timestamps),
            ('Reinforcers (L)', session.timestamps.reinforcer_timestamps),
            ('Lick Onsets (N)', session.timestamps.lick_onset_timestamps),
            ('Lick Offsets (O)', session.timestamps.lick_offset_timestamps),
        ]

        # Write arrays as columns starting at row 5
        start_row = 5
        for col_idx, (header, values) in enumerate(arrays, start=1):
            # Header
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

            # Values
            for row_offset, value in enumerate(values, start=1):
                ws.cell(row=start_row + row_offset, column=col_idx, value=value)

            # Set column width
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        # Freeze header row
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
