"""
Excel exporter for STAR Analyzer V3.
Multi-sheet export with All Sessions + per-subject tabs.
"""

from pathlib import Path
from typing import List, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from core.session_manager import ParsedSession


class ExcelExporter:
    """Exports parsed sessions to multi-sheet Excel files."""

    # Column definitions
    COLUMNS = [
        ('Subject', 10),
        ('Date', 12),
        ('Time', 10),
        ('Experiment', 15),
        ('Stage', 12),
        ('Active', 10),
        ('Inactive', 10),
        ('Licks', 10),
        ('Reinforcers', 12),
        ('Discrim %', 12),
        ('Duration (min)', 14),
        ('Pass', 8),
        ('MSN', 25),
    ]

    # Styles
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF')
    PASS_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    PARTIAL_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    FAIL_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def export_sessions(self, sessions: List[ParsedSession], filepath: Path) -> None:
        """
        Export sessions to Excel with multiple sheets.

        Creates:
        - "All Sessions" sheet with all data
        - Per-subject sheets (e.g., "Subject 1", "Subject 2")
        """
        if not sessions:
            raise ValueError("No sessions to export")

        wb = Workbook()

        # Sheet 1: All Sessions
        ws_all = wb.active
        ws_all.title = "All Sessions"
        self._write_session_sheet(ws_all, sorted(sessions, key=lambda s: (s.subject_id, s.date)))

        # Group sessions by subject
        by_subject = {}
        for session in sessions:
            by_subject.setdefault(session.subject_id, []).append(session)

        # Create per-subject sheets
        for subject_id in sorted(by_subject.keys(), key=self._sort_subject_key):
            sheet_name = f"Subject {subject_id}"[:31]  # Excel limits sheet names to 31 chars
            ws = wb.create_sheet(title=sheet_name)
            subject_sessions = sorted(by_subject[subject_id], key=lambda s: s.date)
            self._write_session_sheet(ws, subject_sessions)

        # Save workbook
        wb.save(filepath)

    def _sort_subject_key(self, subject_id: str):
        """Sort subjects numerically if possible, otherwise alphabetically."""
        try:
            return (0, int(subject_id))
        except ValueError:
            return (1, subject_id)

    def _write_session_sheet(self, ws, sessions: List[ParsedSession]) -> None:
        """Write session data to a worksheet."""
        # Write header row
        for col_idx, (col_name, col_width) in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        # Write data rows
        for row_idx, session in enumerate(sessions, start=2):
            row_data = self._session_to_row(session)

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.THIN_BORDER

                # Center align most columns
                if col_idx not in [4, 13]:  # Experiment and MSN are left-aligned
                    cell.alignment = Alignment(horizontal='center')

            # Color the Pass column based on value
            pass_cell = ws.cell(row=row_idx, column=12)
            if pass_cell.value == 'Pass':
                pass_cell.fill = self.PASS_FILL
            elif pass_cell.value == 'Partial':
                pass_cell.fill = self.PARTIAL_FILL
            elif pass_cell.value == 'Fail':
                pass_cell.fill = self.FAIL_FILL

        # Freeze header row
        ws.freeze_panes = 'A2'

    def _session_to_row(self, session: ParsedSession) -> list:
        """Convert a ParsedSession to a row of values."""
        # Calculate discrimination percentage
        total_lever = session.active_presses + session.inactive_presses
        discrim_pct = (session.active_presses / total_lever * 100) if total_lever > 0 else 0

        # Determine pass status string
        if session.pass_status == 'partial':
            pass_str = 'Partial'
        elif session.passed:
            pass_str = 'Pass'
        else:
            pass_str = 'Fail'

        # Format date (already a string in ParsedSession)
        date_str = session.date or ''

        # Get time from the underlying MedPCSession
        time_str = ''
        if session.session and session.session.metadata.start_time:
            time_str = session.session.metadata.start_time.strftime('%H:%M:%S')

        # Get experiment and MSN from underlying session
        experiment = session.session.metadata.experiment if session.session else ''
        msn = session.session.metadata.msn if session.session else ''

        # Calculate duration in minutes from session_time scalar
        duration_min = 0
        if session.session and session.session.scalars.session_time:
            duration_min = session.session.scalars.session_time / 60

        return [
            session.subject_id,
            date_str,
            time_str,
            experiment,
            session.stage_name,
            session.active_presses,
            session.inactive_presses,
            session.licks,
            session.reinforcers,
            round(discrim_pct, 1),
            round(duration_min, 1),
            pass_str,
            msn,
        ]


class TrackerExporter:
    """Exports tracker progress data to Excel."""

    def export_tracker(self, tracker, filepath: Path) -> None:
        """
        Export tracker state to Excel.

        Creates a summary sheet with animal progress.
        """
        from core.data_models import AnimalState

        wb = Workbook()
        ws = wb.active
        ws.title = "Progress Summary"

        # Header
        headers = ['Subject', 'Current Stage', 'Session Day', 'Consecutive Passes',
                   'Consecutive Fails', 'Total Sessions', 'Status']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = ExcelExporter.HEADER_FILL
            cell.font = ExcelExporter.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = ExcelExporter.THIN_BORDER

        # Data rows
        row_idx = 2
        for animal_id, animal_state in sorted(tracker.animals.items(),
                                               key=lambda x: ExcelExporter()._sort_subject_key(x[0])):
            ws.cell(row=row_idx, column=1, value=animal_id)
            ws.cell(row=row_idx, column=2, value=animal_state.current_stage.name)
            ws.cell(row=row_idx, column=3, value=animal_state.session_day)
            ws.cell(row=row_idx, column=4, value=animal_state.consecutive_passes)
            ws.cell(row=row_idx, column=5, value=animal_state.consecutive_fails)
            ws.cell(row=row_idx, column=6, value=len(animal_state.history))

            # Status
            if animal_state.consecutive_fails >= 3:
                status = 'At Risk'
            elif animal_state.consecutive_passes >= 2:
                status = 'Ready to Advance'
            else:
                status = 'In Progress'
            ws.cell(row=row_idx, column=7, value=status)

            for col_idx in range(1, 8):
                ws.cell(row=row_idx, column=col_idx).border = ExcelExporter.THIN_BORDER
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center')

            row_idx += 1

        # Set column widths
        widths = [10, 15, 12, 18, 16, 15, 15]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = 'A2'
        wb.save(filepath)
